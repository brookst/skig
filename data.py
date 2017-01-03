#!/usr/bin/env python3
#
"""Read acceleration data"""

from __future__ import print_function
from math import sqrt
from numpy import empty as empty_array
from openpyxl import load_workbook


def from_file(file_name, *args):
    """Extract data from file_name"""
    workbook = load_workbook(file_name)
    return from_workbook(workbook, *args)


def from_workbook(workbook, sheet_name, *args):
    """Extract data from open work book"""
    worksheet = workbook[sheet_name]
    return from_worksheet(worksheet, *args)


def from_worksheet(worksheet):
    """Pick data from given worksheet"""
    g_data = empty_array([worksheet.max_row - 1, 2], dtype=float)
    rows = worksheet.rows
    next(rows)  # Eat up the column headers
    index = 0
    for row in rows:
        values = [cell.value for cell in row]
        # print("%d %s" % (index, str(values)))
        if values[0] is None:
            # print("Got %d rows of data" % index)
            break
        else:
            index += 1
        time = values[0] / 1000000.
        acc_1 = values[3]
        acc_2 = values[4]
        acc_3 = values[5]
        mag = sqrt(acc_1 ** 2 + acc_2 ** 2 + acc_3 ** 2)
        g_data[index - 1, 0] = time
        g_data[index - 1, 1] = mag / 9.81
    g_data.resize(index, 2)
    return g_data


def differentiate(data):
    """Differentiate the first column of data"""
    length, _ = data.shape
    for index in range(length - 1):
        current = data[index, 0]
        following = data[index + 1, 0]
        data[index, 0] = following - current
    data[length - 1, 0] = data[length - 2, 0]
    return data


def demo():
    """Proof of principle"""
    workbook = load_workbook('Acceleration_data.xlsx')
    print("Sheet names: " + str(workbook.get_sheet_names()))

    worksheet = workbook[u'Corviglia 30.12.']
    print("Rows: %d" % worksheet.max_row)
    print(' '.join([str(cell.value) for cell in next(worksheet.rows)]))

    # Print contents to treminal
    for index, row in enumerate(worksheet.rows):
        print(' '.join([str(cell.value) for cell in row]))
        if index > 18:
            break

    # Extract a NumPy array of data
    g_data = from_worksheet(worksheet)
    print(g_data)
    print(g_data.shape)
    differentiate(g_data)

    # And all in one shot:
    g_data = from_file('Acceleration_data.xlsx', 'Corviglia 30.12.')
    print(g_data)
    print(g_data.shape)


if __name__ == '__main__':
    demo()
